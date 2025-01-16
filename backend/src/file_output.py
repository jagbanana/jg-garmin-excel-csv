# src/file_output.py

from pathlib import Path
from typing import List
from datetime import date
import csv
import logging
from garmin_client import GarminMetrics

logger = logging.getLogger(__name__)

class FileOutput:
    """Handles exporting Garmin metrics to CSV/XLSX files"""
    
    HEADERS = [
        'Date', 'Sleep Score', 'Sleep Length', 'Weight', 'Body Fat %',
        'Blood Pressure Systolic', 'Blood Pressure Diastolic',
        'Active Calories', 'Resting Calories', 'Resting Heart Rate',
        'Average Stress', 'Training Status',
        'VO2 Max Running', 'VO2 Max Cycling',
        'Intensity Minutes', 'All Activity Count',
        'Running Activity Count', 'Running Distance',
        'Cycling Activity Count', 'Cycling Distance',
        'Strength Activity Count', 'Strength Duration',
        'Cardio Activity Count', 'Cardio Duration'
    ]

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        if not self.output_dir.exists():
            self.output_dir.mkdir(parents=True)

    def _generate_filename(self, start_date: date, end_date: date, extension: str) -> str:
        """Generate filename based on date range"""
        return f"garmin_export_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.{extension}"

    def _metrics_to_row(self, metric: GarminMetrics) -> List:
        """Convert a GarminMetrics object to a row of data"""
        return [
            metric.date.strftime('%Y-%m-%d'),
            metric.sleep_score,
            metric.sleep_length,
            metric.weight,
            metric.body_fat,
            metric.blood_pressure_systolic,
            metric.blood_pressure_diastolic,
            metric.active_calories,
            metric.resting_calories,
            metric.resting_heart_rate,
            metric.average_stress,
            metric.training_status,
            metric.vo2max_running,
            metric.vo2max_cycling,
            metric.intensity_minutes,
            metric.all_activity_count,
            metric.running_activity_count,
            metric.running_distance,
            metric.cycling_activity_count,
            metric.cycling_distance,
            metric.strength_activity_count,
            metric.strength_duration,
            metric.cardio_activity_count,
            metric.cardio_duration
        ]

    def save_csv(self, metrics: List[GarminMetrics], start_date: date, end_date: date) -> Path:
        """Save metrics to CSV file"""
        try:
            filename = self._generate_filename(start_date, end_date, "csv")
            filepath = self.output_dir / filename
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(self.HEADERS)
                
                for metric in metrics:
                    writer.writerow(self._metrics_to_row(metric))
            
            logger.info(f"Successfully saved CSV file: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error saving CSV file: {str(e)}")
            raise

    def save_xlsx(self, metrics: List[GarminMetrics], start_date: date, end_date: date) -> Path:
        """Save metrics to XLSX file"""
        try:
            # Import here to make xlsx support optional
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            filename = self._generate_filename(start_date, end_date, "xlsx")
            filepath = self.output_dir / filename
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Garmin Data"
            
            # Write headers with styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
            
            for col, header in enumerate(self.HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Write data
            for row, metric in enumerate(metrics, 2):
                for col, value in enumerate(self._metrics_to_row(metric), 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            
            wb.save(filepath)
            logger.info(f"Successfully saved XLSX file: {filepath}")
            return filepath
            
        except ImportError:
            logger.warning("openpyxl not installed, XLSX export not available")
            raise
        except Exception as e:
            logger.error(f"Error saving XLSX file: {str(e)}")
            raise